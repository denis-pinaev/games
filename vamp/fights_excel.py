import openpyxl
import sys
import json
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import pyamf
import datetime
import time
from pyamf.remoting.client import RemotingService

data = {}
all_data = []
persons = {}
sid = 11329

def changeInfo(ltype):
    all_data.append(data[ltype])

def getData(bid):
    global data
    client = RemotingService('http://dracula.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    params1 = {
        "userId":"124520",
        "ServerVersion":sid,
        "SystemID":"1",
        "sig":"d6b9ab24005cf6447de56d505f09b9a9"
    }
    resp = service.dispatch(params1,"AMFService.ClanFightInfoGet",[{"BattleDBID":bid,"IsArchive":True}],False)
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("ClanFightInfo"):
        data = data["ClanFightInfo"]
    if data != {}:
        if data["TargetClanDBID"] == 874730:
            changeInfo("TargetPlayers")
        elif data["ActorClanDBID"] == 874730:
            changeInfo("ActorPlayers")

def getBattlesData():
    data = []
    client = RemotingService('http://dracula.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    params1 = {
        "userId":"124520",
        "ServerVersion":sid,
        "SystemID":"1",
        "sig":"d6b9ab24005cf6447de56d505f09b9a9"
    }
    resp = service.dispatch(params1,"AMFService.ClanFightListGet",[{}],False)
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("ClanWarsList"):
        data = data["ClanWarsList"]
    return data

def saveData():
    name = 'fights'
    try:
        wb = load_workbook(name+'.xlsx')
    except:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "log"
        wb.save(name+'.xlsx')
        wb = load_workbook(name+'.xlsx')
    ws = wb.worksheets[0]

    i = 1
    ws.cell('A1').value = 'Name'
    ws.cell('C1').value = 'Dammage'
    ws.cell('D1').value = 'Health'
    ws.cell('B1').value = 'Ignore'
    for n in persons:
        i = i + 1
        ws.cell('A'+str(i)).value = n
        ws.cell('B'+str(i)).value = persons[n]['total']['Ignore']
        ws.cell('C'+str(i)).value = persons[n]['total']['DmgDealed']
        ws.cell('D'+str(i)).value = persons[n]['total']['DmgTake']
        for b in persons[n]:
            if b == 'total': continue
            dd = persons[n][b]['DmgDealed']
            ws.cell(row = 0, column = int(b)+3).value = "battle"+b
            ws.cell(row = i-1, column = int(b)+3).value = dd

    wb.save(name+'.xlsx')

def convert():
    battle_n = 0
    for battle in all_data:
        battle_n = battle_n + 1
        battle_id = str(battle_n)
        for person in battle:
            name = person["PlayerName"]
            if persons.has_key(name):
                persons[name][battle_id] = person
            else:
                persons[name] = {battle_id:person}
    for n in persons:
        total = persons[n]['total'] = {'DmgDealed':0, 'DmgTake':0, 'Ignore':0}
        for b in persons[n]:
            if b == 'total': continue
            total['DmgDealed'] += persons[n][b]['DmgDealed']
            total['DmgTake'] += persons[n][b]['HealthMax'] - persons[n][b]['HealthCur']
            #total['Stamina'] += persons[n][b]['Stamina']
            if persons[n][b]['Stamina'] > 0:
            	total['Ignore'] += 1

battles = getBattlesData()
i = 0
for battle in battles:
    if battle["Winner"] == 0: continue
    bid = battle["BattleDBID"]
    print bid
    getData(bid)
    i = i + 1
    #if i>5: break

convert()
saveData()

