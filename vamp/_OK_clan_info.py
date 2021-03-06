import openpyxl
import sys
import json
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import pyamf
import datetime
import time
from pyamf.remoting.client import RemotingService



sid = 15446
cid = 985939
cname = ''
params1 = {
        "session":"8498TU6YQa6byTB.Idf2M281wd87xXG3x34ayTC3M73bPUkWO362y1DTR6",
        "userId":"909428586714",
        "ServerVersion":sid,
        "SystemID":"4",
        "sig":"65b4fe27f60a84e1789e64f2e83d7601"
}


    
def getClanList():
    data = {}
    client = RemotingService('http://odnvamp.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    resp = service.dispatch(params1,"AMFService.ClanListGet",[{"IsFight":False}],False)
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("ClanList"):
       data = data["ClanList"]
    return data
    
def getClanData():
    data = []
    client = RemotingService('http://odnvamp.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    resp = service.dispatch(params1,"AMFService.ClanInfoGet",[{"ClanDBID": cid}],False)
    name = 'none'
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("PublicClanInfo"):
        data = data["PublicClanInfo"]
        name = data["Name"]
    if data.has_key("ClanUsers"):
        data = data["ClanUsers"]
    return name, data

def saveData(cname, data):
    name = '_OK_clan_info_'+str(cid)
    try:
        wb = load_workbook(name+'.xlsx')
    except:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "log"
        wb.save(name+'.xlsx')
        wb = load_workbook(name+'.xlsx')
    ws = wb.worksheets[0]
    ws.title = cname
    i = 1
    ws.cell('C1').value = 'Name'
    ws.cell('E1').value = 'Level'
    ws.cell('G1').value = 'Ladder'
    ws.cell('H1').value = 'Win'
    ws.cell('I1').value = 'Kill'
    ws.cell('J1').value = 'Mission'
    for n in data:
        i = i + 1
        ws.cell('A'+str(i)).value = '1'
        ws.cell('B'+str(i)).value = '.'
        ws.cell('C'+str(i)).value = n["Name"]
        ws.cell('D'+str(i)).value = '-'
        ws.cell('E'+str(i)).value = n["Lvl"]
        ws.cell('F'+str(i)).value = ','
        ws.cell('G'+str(i)).value = n["Ladder"]
        ws.cell('H'+str(i)).value = n["WinCount"]
        ws.cell('I'+str(i)).value = n["KillCount"]
        ws.cell('J'+str(i)).value = n["DoMissionCount"]

    wb.save(name+'.xlsx')

def saveClansData(cdata):
    name = '_OK_clan_info_all'
    try:
        wb = load_workbook(name+'.xlsx')
    except:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "log"
        wb.save(name+'.xlsx')
        wb = load_workbook(name+'.xlsx')
    ws = wb.worksheets[0]
    i = 0
    j = 0
    clansN = []
    for c in cdata:
        for n in c:
            
            clansN.append(n)
            ws.cell(row=i, column=j).value = n
            j = j + 1
        break
    for c in cdata:
        i = i + 1
        j = 0
        for n in c:
            ws.cell(row=i, column=j).value = c[n]
            j = j + 1

    wb.save(name+'.xlsx')

if len(sys.argv) > 1:
    cid = int(sys.argv[1])

if len(sys.argv) > 2:
    #cname = sys.argv[2]
    clist = getClanList()
    clist.sort(key=lambda x: x['Rang'], reverse=True)
    f = open('_OK_clan_info', 'w')
    for c in clist:
        try:
            f.write("name: %s, id: %s, rang: %s, created: %s\n" % (c['Name'].encode('utf-8'), str(c['ClanDBID']), str(c['Rang']), str(c['CreationDate'])))
        except: print c['Name'], c['ClanDBID']
    f.close()
    saveClansData(clist)
        
print cid, cname
if cid > 0:
    name, data = getClanData()
    saveData(name, data)

