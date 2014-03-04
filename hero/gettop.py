# -*- coding: utf-8 -*-
import sys
import json
import datetime
import time
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from common import *
from gettop_base import *

print auth

person = 2
person_id = '124520'

if len(sys.argv) > 1:
    person_id = sys.argv[1]
if len(sys.argv) > 2:
    person = int(sys.argv[2])
    
persons = [
              {"pid":"124520","auth":"1e365d477c3207804013abaddbb6a0c4","gid":0,"sid":""},#corc
              {"pid":"102137300","auth":"e78c0aad90f427b06653067a45de6c6b","gid":0,"sid":""},#corcc
              {"pid":"160511757","auth":"6dc2dba90c1cc9d935542aa6a60c6fb4","gid":0,"sid":""},#vasya tanakan
              {"pid":"","auth":"","gid":0,"sid":""},#nononon
              {"pid":"155908147","auth":"38e6c4c6f1a3ca43a78a6c499879ba7e","gid":0,"sid":""}#margo
              
          ]
pid = persons[person]["pid"]
auth = persons[person]["auth"]

init_log("gettop")
ctr = int(random.random()*10000)
def getCTR():
    global ctr
    ctr += 1
    return str(ctr)

def getWorld(pers):
    global sid, gid, service, method
    service = actionCommand
    method = 'friendsGetWorld'
    init_params(nsid=sid, ngid=gid, nservice=service, nmethod=method)
    dataString = '{"friendId":"%s","ctr":%s,"sessionKey":"%s","method":"%s"}' % (pers, getCTR(), sid, method)
    params = createData(method, dataString)
    log("%s:%s %s" % (service, method, json.dumps(params)))
    resp = sendRequest(service, params)
    o = json.loads(resp["data"])
    error = o["error"]
    if error == 0:
        log("getWorld %s done" % pers, True)
    else:
        log(resp["data"], True)
    return o

def getTop():
    global sid, gid, service, method
    service = actionCommand
    method = 'getHonorTop'
    init_params(nsid=sid, ngid=gid, nservice=service, nmethod=method)
    #{"ctr":2,"player":null,"method":"getHonorTop","list":null,"place":0,"sessionKey":"5315aacfc2a822.39435727"}
    dataString = '{"player":null,"list":null,"place":0,"ctr":%s,"sessionKey":"%s","method":"%s"}' % (getCTR(), sid, method)
    params = createData(method, dataString)
    log("%s:%s %s" % (service, method, json.dumps(params)))
    resp = sendRequest(service, params)
    o = json.loads(resp["data"])
    error = o["error"]
    if error == 0:
        log("getHonorTop done", True)
    else:
        log(resp["data"], True)
    return o

def putData(ws, letter, i, data, param, default='-'):
    if data.has_key(param): default = str(data[param])
    ws.cell(letter+str(i)).value = default
    ws.cell(letter+'1').value = param

def saveData(data):
    name = 'top_info'
    #try:
    #    wb = load_workbook(name+'.xlsx')
    #except:
    if True:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "gettop"
        wb.save(name+'.xlsx')
        wb = load_workbook(name+'.xlsx')
    ws = wb.worksheets[0]
    ws.title = 'gettop'
    i = 1
    for n in data:
        i = i + 1
        putData(ws,'A',i,n,'NO',i-1)
        putData(ws,'B',i,n,'name','......')
        putData(ws,'C',i,n,'vk')
        putData(ws,'D',i,n,'rang')
        putData(ws,'E',i,n,'level')
        putData(ws,'F',i,n,'epower')
        putData(ws,'G',i,n,'clan_name')
        putData(ws,'H',i,n,'clan_owner')
        if n.has_key('adInfo'):
            n['adInfo'] = json.loads(n['adInfo'])
            putData(ws,'I',i,n['adInfo'],'currency')

    wb.save(name+'.xlsx')






need_info = True
top_len = 100


if True:
    data, gid, sid = init(pid, auth)
    top = getTop()
    top_list = []
    players = get_base()
    for p in top["players"]:
        o = {'pid':str(p),'vk':'vk.com/id'+str(p),'rang':int(top["players"][p]['rating'])}
        if players.has_key(str(p)): o['name'] = players[str(p)].encode('utf-8')
        top_list.append(o)

    top_list = sorted(top_list, key=lambda x : x['rang'], reverse=True)[:top_len]
    
    if need_info:
        for i in range(top_len):
            f = getWorld(top_list[i]['pid'])
            f = f["friend"]
            for d in f["player"]:
                top_list[i][d] = f["player"][d]
                if f.has_key("clanInfo") and len(f["clanInfo"])>0 and f["clanInfo"].has_key("ownerId"):
                    ci = f["clanInfo"]
                    top_list[i]['clan_name'] = ci['name'].encode('utf-8')
                    top_list[i]['clan_owner'] = ci['ownerId']
                    top_list[i]['clan_info'] = ci

    saveData(top_list)

