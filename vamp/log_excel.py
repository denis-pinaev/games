import openpyxl
import sys
import json
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

name = '1'
if len(sys.argv) > 1:
    name = sys.argv[1]

cell_type = [
    {"cell":'A', "value":"Time"},
    {"cell":'B', "value":"Type"},
    {"cell":'C', "value":"PlayerName"},
    {"cell":'D', "value":"Change"},
    {"cell":'E', "value":"HealthCur"},
    {"cell":'F', "value":"HealthMax"},
    {"cell":'G', "value":"DmgDealed"},
    {"cell":'H', "value":"Ladder"},
    {"cell":'I', "value":"Stamina"},
    {"cell":'J', "value":"IsDie"},
    {"cell":'K', "value":"FighterDBID"},
]

def getCell(value):
    for p in cell_type:
        if p['value'] == value: return p['cell']
    return None

try:
    wb = load_workbook(name+'.xlsx')
except:
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "log"
    for line in cell_type:
        #line = cell_type[p]
        ws.cell(line["cell"]+'1').value = line["value"]
    wb.save(name+'.xlsx')
    wb = load_workbook(name+'.xlsx')

ws = wb.worksheets[0]

f = open(name, 'r')
data = f.read()
f.close()
data = data.split('\n')
ds = data[len(data)-3]

def insertLine(ds, i):
    ind = ds.index(']')
    time = ds[1:ind]
    ds = ds[ind+1:]
    ind = ds.index(']')
    ptype = ds[1:ind]
    ds = '['+ds[ind+2:]+']'
    ds = json.loads(ds)
    for param in ds:
        if param.has_key('changed'):
            p = param['changed']
            cell = getCell(p)
            ws.cell(cell+str(i)).value = param['new']
            cell = getCell("Change")
            ws.cell(cell+str(i)).value = param['changed']+": "+param['new']+"/"+param['old']+"("+str(int(param['new'])-int(param['old']))+")"
            cell = getCell("Time")
            ws.cell(cell+str(i)).value = time
            cell = getCell("Type")
            ws.cell(cell+str(i)).value = ptype
        else:
            for p in param:
                cell = getCell(p)
                ws.cell(cell+str(i)).value = param[p]

i = 2
for ds in data:
    try:
        insertLine(ds, i)
        i = i + 1
    except: None
    print i
wb.save(name+'.xlsx')