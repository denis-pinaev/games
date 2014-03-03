import openpyxl
import sys
import json
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import pyamf
import sys
import json
import datetime
import time
from pyamf.remoting.client import RemotingService

data = {}
bid = 40252
sid = 22442
name = 'fights_all'
log_start = 2

if len(sys.argv) > 1:
    bid = int(sys.argv[1])

def saveData(i, params):
    try:
        wb = load_workbook(name+'.xlsx')
    except:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "log"
        wb.save(name+'.xlsx')
        wb = load_workbook(name+'.xlsx')
    ws = wb.worksheets[0]

    global bid
    global log_start
    i = log_start
    while True:
        cbid = ws.cell('I'+str(i)).value
        if not cbid: break
        if int(cbid) == bid:
            if not ws.cell('C'+str(i)).value:# and params[2]:
                break
            else:
                del ws
                del wb
                return
        i += 1
    log_start = i
    ws.cell('A1').value = 'Attacker'
    ws.cell('B1').value = 'Target'
    ws.cell('C1').value = 'Winner'
    ws.cell('D1').value = 'Rang'
    ws.cell('E1').value = 'A Dammage'
    ws.cell('F1').value = 'T Dammage'
    ws.cell('G1').value = 'A Count'
    ws.cell('H1').value = 'T Count'
    ws.cell('I1').value = 'Battle ID'
    for j in range(len(params)):
        letter = chr(ord('A') + j)
        ws.cell(letter+str(i)).value = params[j]
    wb.save(name+'.xlsx')
    del ws
    del wb

def log(s):
    #s = "[" + datetime.datetime.now().strftime('%d.%m %H:%M:%S')+"]" + s
    fatype = "f_log_total_all"
    try:
        tfile = open(fatype, "a")
        tfile.write(s.encode('utf-8')+'\n')
        tfile.close()
    except:
        tfile = open(fatype, "w")
        tfile.write(s.encode('utf-8')+'\n')
        tfile.close()

def getClanList():
    data = {}
    client = RemotingService('http://dracula.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    params1 = {
        "userId":"124520",
        "ServerVersion":sid,
        "SystemID":"1",
        "sig":"d6b9ab24005cf6447de56d505f09b9a9"
    }
    resp = service.dispatch(params1,"AMFService.ClanListGet",[{"IsFight":False}],False)
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("ClanList"):
       data = data["ClanList"]
    return data

def getBid():
    global bid
    bf = open(name, 'r')
    bid = int(bf.read())
    bf.close()
    return bid

def saveBid(bid):
    bf = open(name, 'w')
    bf.write(str(bid))
    bf.close()

def getData(arch=False):
    data = {}
    client = RemotingService('http://dracula.irq.ru/Gateway.aspx')
    service = client.getService('MafiaAMF.AMFService')
    params1 = {
        "userId":"124520",
        "ServerVersion":sid,
        "SystemID":"1",
        "sig":"d6b9ab24005cf6447de56d505f09b9a9"
    }
    resp = service.dispatch(params1,"AMFService.ClanFightInfoGet",[{"BattleDBID":bid,"IsArchive":arch}],False)
    to_end = False
    if resp.has_key("data"):
        data = resp["data"]
    if data.has_key("Answer"):
        data = data["Answer"]
    if data.has_key("OtherData"):
        data = data["OtherData"]
    if data.has_key("ClanFightInfo"):
       data = data["ClanFightInfo"]
       to_end = True
    return data, to_end

clans = getClanList()
last_bid = getBid()
last_not_finished = 9999999
if clans:
    cn = []
    clans.sort(key=lambda x: x['Rang'], reverse=True)
    for i in range(30):
        c = clans[i]
        s = ''
        for k in c:
            s+='%s:%s, ' % (k, c[k])
        cn.append(c['Name'])
        log(s)
    bcnt = 1
    while True:
        b_act = 'Active'
        data, to_end = getData()
        if not data:
            b_act = 'Finished'
            data, to_end = getData(True)
        if data:
            last_bid = bid
            A = data["ActorClanName"]
            T = data["TargetClanName"]
            if A in cn or T in cn:
                bcnt += 1
                a_cnt = str(len(data["ActorPlayers"]))
                t_cnt = str(len(data["TargetPlayers"]))
                a_dmg = 0
                for p in data["ActorPlayers"]:
                    a_dmg += p["DmgDealed"]
                a_dmg = str(a_dmg)
                t_dmg = 0
                for p in data["TargetPlayers"]:
                    t_dmg += p["DmgDealed"]
                t_dmg = str(t_dmg)
                s = "A: %s, T: %s, id:%s. Dammage: A=%s(%s), T=%s(%s)" % (A, T, str(bid), a_cnt, a_dmg, t_cnt, t_dmg)
                winner = ''
                rang = ''
                if b_act == "Finished":
                    winner = data["ActorClanName"] if data["ActorClanDBID"] == data["Winner"] else data["TargetClanName"]
                    rang = str(data["LadderPtsDelta"])
                    s = "%s. Winner: %s, rang: %s" % (s, winner, rang)
                elif bid < last_not_finished:
                    saveBid(bid)
                    last_not_finished = bid
                saveData(bcnt, [A, T, winner, rang, a_dmg, t_dmg, a_cnt, t_cnt, bid])
                print s
                log(s)
        else:
            if bid - last_bid > 20:
                print 'End fights:', last_bid
                bid = last_bid
                break
        if to_end:
            bid += 1
        else:
            print 'end_fights_error', last_bid
            time.sleep(30*60)
print 'global END'